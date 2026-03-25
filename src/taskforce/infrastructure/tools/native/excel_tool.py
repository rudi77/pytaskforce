"""Excel spreadsheet manipulation tool.

Provides reading, writing, and editing capabilities for Microsoft Excel
(.xlsx) files.  Requires the ``openpyxl`` library (install via
``uv sync --extra office``).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import structlog

from taskforce.core.interfaces.tools import ApprovalRiskLevel
from taskforce.infrastructure.tools.base_tool import BaseTool

logger = structlog.get_logger(__name__)


def _require_openpyxl() -> Any:
    """Lazily import openpyxl and raise a clear error if missing."""
    try:
        import openpyxl

        return openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: uv sync --extra office"
        )


class ExcelTool(BaseTool):
    """Read, create, and edit Microsoft Excel (.xlsx) spreadsheets."""

    tool_name = "excel"
    tool_description = (
        "Manipulate Microsoft Excel (.xlsx) files. "
        "Supports reading sheets, creating workbooks, writing cells/rows, "
        "adding sheets, and getting workbook info."
    )
    tool_parameters_schema = {
        "type": "object",
        "properties": {
            "action": {
                "type": "string",
                "enum": [
                    "read",
                    "create",
                    "write_cells",
                    "append_rows",
                    "add_sheet",
                    "info",
                ],
                "description": (
                    "Action to perform: "
                    "'read' reads data from a sheet, "
                    "'create' creates a new workbook, "
                    "'write_cells' writes values to specific cells, "
                    "'append_rows' appends rows to a sheet, "
                    "'add_sheet' adds a new sheet, "
                    "'info' returns workbook metadata."
                ),
            },
            "path": {
                "type": "string",
                "description": "Path to the .xlsx file.",
            },
            "sheet": {
                "type": "string",
                "description": "Sheet name (default: active sheet).",
            },
            "cells": {
                "type": "object",
                "description": (
                    "Cell values to write as {cell_ref: value} mapping, "
                    "e.g. {'A1': 'Name', 'B1': 'Age', 'A2': 'Alice', 'B2': 30}."
                ),
            },
            "rows": {
                "type": "array",
                "items": {
                    "type": "array",
                },
                "description": "Rows to append, each row a list of values.",
            },
            "headers": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Header row for create action.",
            },
            "max_rows": {
                "type": "integer",
                "description": "Maximum number of rows to read (default: 1000).",
            },
        },
        "required": ["action", "path"],
    }

    tool_requires_approval = True
    tool_approval_risk_level = ApprovalRiskLevel.MEDIUM
    tool_supports_parallelism = True

    async def _execute(self, **kwargs: Any) -> dict[str, Any]:
        """Dispatch to the requested action handler."""
        action = kwargs["action"]
        handler = {
            "read": self._read,
            "create": self._create,
            "write_cells": self._write_cells,
            "append_rows": self._append_rows,
            "add_sheet": self._add_sheet,
            "info": self._info,
        }.get(action)

        if handler is None:
            return {"success": False, "error": f"Unknown action: {action}"}

        return await handler(**kwargs)

    async def _read(self, **kwargs: Any) -> dict[str, Any]:
        """Read data from a sheet."""
        openpyxl = _require_openpyxl()
        path = kwargs["path"]
        sheet_name = kwargs.get("sheet")
        max_rows = kwargs.get("max_rows", 1000)

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > max_rows:
                break
            rows_data.append([self._cell_to_str(c) for c in row])

        wb.close()

        return {
            "success": True,
            "sheet": ws.title,
            "rows": rows_data,
            "row_count": len(rows_data),
            "column_count": len(rows_data[0]) if rows_data else 0,
        }

    async def _create(self, **kwargs: Any) -> dict[str, Any]:
        """Create a new .xlsx workbook with optional headers and data."""
        openpyxl = _require_openpyxl()
        path = kwargs["path"]
        sheet_name = kwargs.get("sheet", "Sheet1")
        headers = kwargs.get("headers")
        rows = kwargs.get("rows", [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if headers:
            ws.append(headers)
        for row in rows:
            ws.append(row)

        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        wb.close()

        logger.info("excel.created", path=path, sheet=sheet_name)
        return {"success": True, "path": path, "message": f"Workbook created at {path}"}

    async def _write_cells(self, **kwargs: Any) -> dict[str, Any]:
        """Write values to specific cells."""
        openpyxl = _require_openpyxl()
        path = kwargs["path"]
        sheet_name = kwargs.get("sheet")
        cells = kwargs.get("cells", {})

        if not cells:
            return {"success": False, "error": "Parameter 'cells' is required for write_cells"}

        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        written = 0
        for cell_ref, value in cells.items():
            ws[cell_ref] = value
            written += 1

        wb.save(path)
        wb.close()

        logger.info("excel.cells_written", path=path, count=written)
        return {
            "success": True,
            "path": path,
            "cells_written": written,
            "message": f"Wrote {written} cell(s)",
        }

    async def _append_rows(self, **kwargs: Any) -> dict[str, Any]:
        """Append rows to a sheet."""
        openpyxl = _require_openpyxl()
        path = kwargs["path"]
        sheet_name = kwargs.get("sheet")
        rows = kwargs.get("rows", [])

        if not rows:
            return {"success": False, "error": "Parameter 'rows' is required for append_rows"}

        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        for row in rows:
            ws.append(row)

        wb.save(path)
        wb.close()

        logger.info("excel.rows_appended", path=path, count=len(rows))
        return {
            "success": True,
            "path": path,
            "rows_appended": len(rows),
            "message": f"Appended {len(rows)} row(s)",
        }

    async def _add_sheet(self, **kwargs: Any) -> dict[str, Any]:
        """Add a new sheet to an existing workbook."""
        openpyxl = _require_openpyxl()
        path = kwargs["path"]
        sheet_name = kwargs.get("sheet", "New Sheet")

        wb = openpyxl.load_workbook(path)
        if sheet_name in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' already exists"}

        wb.create_sheet(title=sheet_name)
        wb.save(path)
        wb.close()

        logger.info("excel.sheet_added", path=path, sheet=sheet_name)
        return {
            "success": True,
            "path": path,
            "sheet": sheet_name,
            "message": f"Sheet '{sheet_name}' added",
        }

    async def _info(self, **kwargs: Any) -> dict[str, Any]:
        """Return workbook metadata."""
        openpyxl = _require_openpyxl()
        path = kwargs["path"]

        wb = openpyxl.load_workbook(path, read_only=True)

        sheets_info = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets_info.append({
                "name": name,
                "min_row": ws.min_row,
                "max_row": ws.max_row,
                "min_column": ws.min_column,
                "max_column": ws.max_column,
            })

        props = wb.properties
        result = {
            "success": True,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_info,
            "author": props.creator or "",
            "title": props.title or "",
            "created": str(props.created) if props.created else "",
            "modified": str(props.modified) if props.modified else "",
        }
        wb.close()
        return result

    @staticmethod
    def _cell_to_str(value: Any) -> Any:
        """Convert a cell value to a JSON-serializable form."""
        if value is None:
            return None
        return value
